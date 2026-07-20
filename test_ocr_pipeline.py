"""
端到端管線測試（mock 掉 Tesseract，不需安裝 OCR 引擎）。
驗證：process_image -> 解析 -> excel_writer.build 產出可用 xlsx。
執行：python test_ocr_pipeline.py
"""
import io
from unittest import mock

import ocr
import excel_writer

MRZ_TEXT = (
    "P<GBRSTEPHENS<<JOHN<<MR<MICHAEL" + "<" * 12 + "\n"
    "GBN123456<GBR8501018M2501018" + "<" * 15 + "0"
)


def fake_tesseract(img, config=""):
    return MRZ_TEXT


def fake_prep_mrz(img, upscale_to):
    return img


def fake_open(bio):
    img = mock.MagicMock()
    img.size = (1200, 800)
    img.convert.return_value = img
    return img


with mock.patch.object(ocr.pytesseract, "image_to_string", fake_tesseract), \
     mock.patch.object(ocr, "_prep_mrz", fake_prep_mrz), \
     mock.patch.object(ocr.Image, "open", fake_open):
    res = ocr.process_image(b"fake-bytes")

print("OCR 結果：", res)
assert res and res["parsed"]["doc_number"] == "GBN123456", "MRZ 解析失敗"

rec = {"doc_type": "護照", "scan_time": "2026-07-16 14:00:00", "parsed": res["parsed"]}
path = excel_writer.build([rec])
print("Excel 已產出：", path)

import openpyxl  # noqa: E402

wb = openpyxl.load_workbook(path)
ws = wb.active
print("標頭：", [c.value for c in ws[1]])
print("資料：", [c.value for c in ws[2]])
assert ws[2][3].value == "GBN123456", "Excel 欄位錯誤"
print("\n✅ 端到端管線通過")
