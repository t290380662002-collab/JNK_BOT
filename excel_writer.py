"""
用 openpyxl 產生證件資料 Excel。
每筆記錄（record）結構：
  {
    "doc_type": "護照" | "港澳通行證" | "回鄉證" | "台胞證",
    "scan_time": "2026-07-16 14:30:00",
    "parsed": { last_name, first_name, doc_number, nationality,
                date_of_birth, sex, expiry_date, issuer, doc_type_guess }
  }
"""
import os
import tempfile
from datetime import datetime

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill

# (record 層級 key 或 parsed 層級 key, 中文欄位名)
FIELDS = [
    ("doc_type", "證件類型"),
    ("last_name", "英文姓"),
    ("first_name", "英文名"),
    ("doc_number", "證件號碼"),
    ("nationality", "國籍/地區"),
    ("date_of_birth", "出生日期"),
    ("sex", "性別"),
    ("expiry_date", "有效期限"),
    ("issuer", "發證代碼"),
    ("scan_time", "掃描時間"),
]

HEADER_FILL = PatternFill("solid", fgColor="2F5496")
HEADER_FONT = Font(bold=True, color="FFFFFF")
CENTER = Alignment(horizontal="center", vertical="center")


def build(records: list, path: str | None = None) -> str:
    wb = Workbook()
    ws = wb.active
    ws.title = "證件資料"

    # 標頭
    for col, (_, header) in enumerate(FIELDS, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER

    # 資料列
    for row_idx, rec in enumerate(records, start=2):
        parsed = rec.get("parsed", {})
        for col, (key, _) in enumerate(FIELDS, start=1):
            if key in ("doc_type", "scan_time"):
                value = rec.get(key, "")
            else:
                value = parsed.get(key, "")
            ws.cell(row=row_idx, column=col, value=value)

    # 自動欄寬
    for col_cells in ws.columns:
        width = 10
        for c in col_cells:
            if c.value is not None:
                width = max(width, len(str(c.value)))
        letter = col_cells[0].column_letter
        ws.column_dimensions[letter].width = min(width + 2, 40)

    ws.freeze_panes = "A2"

    if path is None:
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(tempfile.gettempdir(), f"docscan_{ts}.xlsx")

    wb.save(path)
    return path
