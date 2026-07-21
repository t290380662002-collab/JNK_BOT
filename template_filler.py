# -*- coding: utf-8 -*-
"""
把客人資料填入六間酒店的訂房單模板。

對外主函式：
    fill(hotel_key, records)         # records: 掃描記錄清單（photo OCR）
    fill_manual(hotel_key, booking)  # booking: 文字訂房 dict（/book 指令）

record（掃描）結構：
    {"doc_type":..., "parsed": {last_name, first_name, doc_number, date_of_birth, ...}}
booking（文字訂房）結構：
    {
      "guests": ["中文姓名", ...],
      "en_names": ["ENG,NAME", ...] (optional, 與 guests 對齊),
      "check_in": "2026-07-21", "check_out": "2026-07-23",
      "room_count": 1, "pax": 1,
      "smoking": True / False / None,
      "doc_numbers": [...], "dobs": [...] (optional, 與 guests 對齊),
    }

行為（依用戶確認）：
  · 有清單表的酒店：清單表填全部客人（掃描時中文姓名留空手動補；文字訂房則填入中文姓名）
  · 每位客人各產生一張訂房單主表格（label-driven 填入）
  · 房型：若輸入含明確房型代碼（如 TC）或完整中文房型名，自動在 RM TYPE 區域勾選對應選項
  · 吸菸：True/False 會填入訂房單主表格「特別要求 Special request」欄（吸菸/禁煙）
"""
import os
import re
import tempfile
from datetime import datetime, date

import openpyxl
from openpyxl.styles import Alignment, Border, Side

import hotel_templates as HT
from timeutil import taipei_now


# ---------------------------------------------------------------------------
# 值格式化
# ---------------------------------------------------------------------------
def _en_name(parsed: dict) -> str:
    """英文姓名 -> 'LASTNAME,FIRSTNAME'（比照範例格式）。"""
    last = (parsed.get("last_name") or "").strip().upper()
    first = (parsed.get("first_name") or "").strip().upper()
    if last and first:
        return f"{last},{first}"
    return last or first


def _date_value(s):
    """'YYYY-MM-DD' / 'MM-DD' -> datetime（Excel 會格式化）；失敗回原字串。"""
    if not s or not str(s).strip():
        return ""
    s = str(s).strip().replace("/", "-")
    for fmt in ("%Y-%m-%d", "%m-%d"):
        try:
            d = datetime.strptime(s, fmt)
            if fmt == "%m-%d":
                d = d.replace(year=taipei_now().year)
            return d
        except ValueError:
            continue
    return s


def _doc_value(parsed: dict) -> str:
    return (parsed.get("doc_number") or "").strip().upper()


def _split_name(zh: str):
    """中文姓名拆 姓/名：有 '-' 取第一段為姓，否則取首字為姓。"""
    zh = (zh or "").strip()
    if not zh:
        return "", ""
    if "-" in zh:
        parts = [p for p in zh.split("-") if p]
        if parts:
            return parts[0], "-".join(parts[1:])
    return zh[0], zh[1:]


def _split_en_name(en: str):
    """英文姓名拆 姓/名：'WU，JUN' -> ('WU','JUN')；'STEPHENS,JOHN A' -> ('STEPHENS','JOHN A')。"""
    en = (en or "").replace("，", ",").replace("、", ",").strip()
    parts = [p.strip() for p in re.split(r"[,/\s]+", en) if p.strip()]
    if not parts:
        return "", ""
    return parts[0], (" ".join(parts[1:]) if len(parts) > 1 else "")


# ---------------------------------------------------------------------------
# 標準化：record / booking -> 內部 booking dict
# ---------------------------------------------------------------------------
def _normalize_record(rec: dict) -> dict:
    b = {k: "" for k in ("surname", "firstname", "docnum", "dob",
                          "checkin", "checkout", "pax", "rooms",
                          "zh", "en", "room", "smoking", "group", "agent")}
    parsed = rec.get("parsed")
    if parsed:
        b["surname"] = (parsed.get("last_name") or "").strip().upper()
        b["firstname"] = (parsed.get("first_name") or "").strip().upper()
        b["docnum"] = _doc_value(parsed)
        b["dob"] = _date_value(parsed.get("date_of_birth"))
        b["en"] = _en_name(parsed)
        zh = (parsed.get("zh_name") or "").strip()
        en = (parsed.get("en_name") or "").strip()
        if zh:
            b["zh"] = zh
        if en:
            b["en"] = en.upper()
            s, f = _split_en_name(en)
            b["surname"], b["firstname"] = s.upper(), f.upper()
        elif zh:
            # 無英文時用中文拆 姓/名
            s, f = _split_name(zh)
            b["surname"], b["firstname"] = s, f
        # 掃描無中文姓名/房型 -> 留空手動補
    manual = rec.get("manual") or rec.get("booking")
    if manual:
        en = (manual.get("en_name") or "").strip()
        zh = (manual.get("zh_name") or "").strip()
        if en:
            b["en"] = en.upper()
            s, f = _split_en_name(en)
            b["surname"], b["firstname"] = s.upper(), f.upper()
        if zh:
            b["zh"] = zh
            if not en:                      # 無英文時才用中文拆 姓/名
                s, f = _split_name(zh)
                b["surname"], b["firstname"] = s, f
        if manual.get("doc_number"):
            b["docnum"] = str(manual["doc_number"]).strip().upper()
        if manual.get("dob"):
            b["dob"] = _date_value(manual["dob"])
        if manual.get("check_in"):
            b["checkin"] = _date_value(manual["check_in"])
        if manual.get("check_out"):
            b["checkout"] = _date_value(manual["check_out"])
        if manual.get("room_count"):
            b["rooms"] = manual["room_count"]
        if manual.get("pax"):
            b["pax"] = manual["pax"]
        if manual.get("smoking") is not None:
            b["smoking"] = manual["smoking"]
        # 群組/代理：文字訂房帶入清單表
        b["group"] = manual.get("group") or manual.get("wechat") or ""
        b["agent"] = manual.get("agent") or manual.get("booker") or ""
        # 房型：若輸入明確則自動勾選（代碼優先）
        if manual.get("room_type"):
            b["room_type"] = manual["room_type"]
    return b


def _bookings_from_manual(booking: dict) -> list:
    """文字訂房 dict -> 標準化 booking 清單（每位客人一筆，共用日期/房數/吸煙）。

    booking["guests"] 為 dict 清單，每筆含 zh_name/en_name/doc_number/dob（部分可空）。
    """
    guests = booking.get("guests") or []
    out = []
    for g in guests:
        if isinstance(g, str):
            g = {"zh_name": g}
        m = {
            "zh_name": g.get("zh_name"),
            "en_name": g.get("en_name"),
            "doc_number": g.get("doc_number"),
            "dob": g.get("dob"),
            "check_in": booking.get("check_in"),
            "check_out": booking.get("check_out"),
            "room_count": booking.get("room_count"),
            "room_type": booking.get("room_type"),
            "pax": booking.get("pax"),
            "smoking": booking.get("smoking"),
            "wechat": booking.get("wechat"),
            "booker": booking.get("booker"),
            "group": booking.get("group"),
            "agent": booking.get("agent"),
        }
        out.append(_normalize_record({"manual": m}))
    return out


# ---------------------------------------------------------------------------
# label-driven：找標籤右側輸入格
# ---------------------------------------------------------------------------
def _norm(s) -> str:
    return str(s).replace(" ", "").replace("\u3000", "").replace("\n", "")


def _merged_range_of(ws, row, col):
    for mr in ws.merged_cells.ranges:
        if mr.min_row <= row <= mr.max_row and mr.min_col <= col <= mr.max_col:
            return mr
    return None


def _find_label(ws, subs):
    for row in ws.iter_rows():
        for c in row:
            if c.value and isinstance(c.value, str):
                v = _norm(c.value)
                for sub in subs:
                    if _norm(sub) in v:
                        return c
    return None


def _find_date_label(ws):
    """專找獨立的 'Date:' 標籤（避開 C/I Date、C/O Date）。"""
    for row in ws.iter_rows():
        for c in row:
            if c.value and isinstance(c.value, str):
                v = c.value.strip()
                if v.startswith("Date:") and not re.search(r"C\s*[/_]\s*[IO]|入住|退房", v, re.I):
                    return c
    return None


def _input_cell(ws, label_cell):
    """回傳標籤右側輸入格（若在合併範圍則取左上角）。"""
    mr = _merged_range_of(ws, label_cell.row, label_cell.column)
    if mr:
        trow, tcol = mr.min_row, mr.max_col + 1
    else:
        trow, tcol = label_cell.row, label_cell.column + 1
    tmr = _merged_range_of(ws, trow, tcol)
    if tmr:
        trow, tcol = tmr.min_row, tmr.min_col
    return ws.cell(row=trow, column=tcol)


_DOB_FMT = "yyyy/mm/dd"      # 出生日期固定顯示（補零，不隨區域變動）
_CHECK_FMT = "yyyy/mm/dd"    # 入住/退房日期：用戶確認為 2026/07/29 格式


def _clean_special_label(s):
    """把模板裡的『特別要求 Special request  ：』規整成『特別要求 Special request』，
    移除尾端空白與多餘冒號（連同冒號前的空白），便於後面接『：吸菸/禁煙』。"""
    if not s:
        return "特別要求 Special request"
    s = str(s)
    # 先去掉尾端（空白 + 冒號 + 空白），再摺疊內部多餘空白為單一空白
    s = re.sub(r"[\s\u3000]*[：:]+[\s\u3000]*$", "", s)
    s = re.sub(r"[\s\u3000]+", " ", s).strip()
    return s


# ---------------------------------------------------------------------------
# 房型自動勾選
# ---------------------------------------------------------------------------
_ROOM_CODE_RE = re.compile(r"\(([A-Z0-9/]+)\)")


def _pinyin(text: str) -> str:
    """將中文轉為無調號拼音，用於簡繁/異體字比對。"""
    try:
        from pypinyin import lazy_pinyin
        return "".join(lazy_pinyin(text or "", style=0)).lower()
    except Exception:  # noqa: BLE001
        return (text or "").lower()


def _char_similarity(a: str, b: str) -> float:
    """字元層級相似度（簡繁通用）：相同字元或同拼音即視為匹配。"""
    if not a or not b:
        return 0.0
    pa, pb = _pinyin(a), _pinyin(b)
    if pa == pb:
        return 1.0
    # 簡單的共同字元比例
    set_a = set(a)
    set_b = set(b)
    inter = set_a & set_b
    union = set_a | set_b
    if not union:
        return 0.0
    return len(inter) / len(union)


def _find_room_type_cells(ws):
    """找出訂房單主表格中 RM TYPE 區域的所有帶勾選框選項格。"""
    cells = []
    for row in ws.iter_rows(min_row=1, max_row=80):
        for c in row:
            if c.value and isinstance(c.value, str):
                v = c.value
                # 勾選框格式：(      ) 中文房型 (CODE)
                # 已勾選格式：(✓) 中文房型 (CODE)
                if _ROOM_CODE_RE.search(v) and re.search(r"^\([\s✓]*\)", v):
                    cells.append(c)
    return cells


def _room_option_name(option_text: str) -> str:
    """從選項文字抽取出中文房型名稱，例如 '( ) 維多利亞套房 (TC)' -> '維多利亞套房'。"""
    # 去掉開頭勾選框
    body = re.sub(r"^\([\s✓]*\)\s*", "", option_text)
    # 去掉尾端代碼括號
    body = _ROOM_CODE_RE.sub("", body).strip()
    return body


def _match_room_type(room_type_input: str, option_text: str) -> int:
    """回傳房型輸入與模板選項的匹配分數（0~100）。"""
    if not room_type_input or not option_text:
        return 0
    rt = str(room_type_input).lower().replace("，", ",").replace("、", ",")
    opt = option_text.lower()
    opt_name = _room_option_name(option_text).lower()

    # 1) 房型代碼完全匹配（如 TC、KC、DBK1）
    code_match = _ROOM_CODE_RE.findall(option_text)
    for code in code_match:
        # 代碼需以獨立片段出現在輸入中
        if re.search(r"(?:^|[\s,/（(])" + re.escape(code.lower()) + r"(?:$|[\s,/）)])", rt):
            return 100

    # 2) 中文房型名稱完全包含
    if opt_name and opt_name in rt:
        return 95
    if opt_name and rt in opt_name:
        # 若輸入僅為過泛詞（如只有「大床」），視為不夠明確，降低分數
        if rt in ("大床", "雙床", "客房", "套房", "房", "大", "雙"):
            return 50
        return 90

    # 2.5) 簡繁/異體字拼音匹配（如 维多利亚套房 vs 維多利亞套房）
    if opt_name and _pinyin(rt) == _pinyin(opt_name):
        return 88

    # 3) 關鍵字匹配（按共同詞數與長度加權）
    # 先清理出一些有意義的關鍵字（去掉「房」「套房」等過泛詞）
    def tokens(s):
        s = re.sub(r"[\(\)（）,，、/\\\s]+", " ", s)
        return [t for t in s.split() if len(t) >= 2]

    opt_tokens = tokens(opt_name)
    rt_tokens = tokens(rt)
    if not opt_tokens or not rt_tokens:
        return 0
    score = 0
    for t in opt_tokens:
        # 套房/客房/房等泛詞權重較低
        weight = 15 if t in ("套房", "客房", "房") else 30
        if t in rt_tokens:
            score += weight
        # 部分匹配（如「維多利亞」命中「维多利亚」）
        elif any(t in r or r in t for r in rt_tokens):
            score += weight // 2
    return min(score, 89)  # 低於直接匹配門檻


def _reset_room_type_cells(ws, orig_room_cells):
    """強制把 RM TYPE 區域的勾選框還原為原始未勾選文字。"""
    if not orig_room_cells:
        return
    for (row, col), orig_val in orig_room_cells.items():
        ws.cell(row=row, column=col).value = orig_val


def _fill_room_type(ws, room_type):
    """房型自動勾選（已停用）：依用戶要求，RM TYPE 留空由人工選擇。"""
    # 保留函式簽名向後相容，但直接返回不做任何勾選
    return


def _fill_special_request(ws, lc, value, orig_label):
    """特別要求：『特別要求 Special request：吸菸』顯示在「原本標籤合併格」內。

    - 保留模板原來的標籤合併範圍（如 L14:O14），不向右擴張到輸入格，避免超出原框。
    - 把整串文字寫入標籤格左上角，開啟自動換行 + 縮小字型，並水平垂直置中，
      讓文字能完整落在原框框內。
    - 文字基底取自原始模板標籤（orig_label），避免複製表累積上次的值。
    """
    label_base = _clean_special_label(orig_label if orig_label else lc.value)
    full = f"{label_base}：{value}"

    # 1) 取得現有合併範圍；若原標籤未合併，則只與右側一格合併
    mr = _merged_range_of(ws, lc.row, lc.column)
    if mr:
        start_col, end_col = mr.min_col, mr.max_col
    else:
        start_col = lc.column
        end_col = min(lc.column + 1, 20)
        ws.merge_cells(start_row=lc.row, start_column=start_col,
                       end_row=lc.row, end_column=end_col)

    # 2) 清除右側原輸入格內容，避免舊值殘留
    if mr:
        for c in range(mr.max_col + 1, 21):
            adj = ws.cell(row=lc.row, column=c)
            if adj.value is not None:
                adj.value = None

    # 3) 寫入整串文字並設定對齊/換行/縮小
    top = ws.cell(row=lc.row, column=start_col)
    top.value = full
    top.alignment = Alignment(
        horizontal="center", vertical="center",
        wrap_text=True, shrink_to_fit=True,
    )


def _fill_form_sheet(ws, b: dict, orig_special_label=None, orig_room_cells=None):
    """label-driven 填一張訂房單主表格（只填有值的欄，避免清掉模板）。"""
    # 房型欄強制重置為原始未勾選狀態（雙保險，避免舊模板/舊資料殘留）
    _reset_room_type_cells(ws, orig_room_cells)
    # 吸菸 -> 特別要求欄；True=吸菸，False=禁煙，None=不填
    special = ""
    if b.get("smoking") is True:
        special = "吸菸"
    elif b.get("smoking") is False:
        special = "禁煙"
    values = {
        "surname": b.get("surname", ""),
        "firstname": b.get("firstname", ""),
        "docnum": b.get("docnum", ""),
        "dob": b.get("dob", ""),
        "checkin": b.get("checkin", ""),
        "checkout": b.get("checkout", ""),
        "pax": b.get("pax", ""),
        "rooms": b.get("rooms", ""),
        "special_request": special,
        "date": taipei_now().date(),
    }
    for field, subs in HT.FORM_LABELS.items():
        if field == "date":
            # Date: 欄位需避免誤匹配 C/I Date / C/O Date，改由專用搜尋
            lc = _find_date_label(ws)
        else:
            lc = _find_label(ws, subs)
        if lc is None:
            continue
        v = values.get(field, "")
        if v == "" or v is None:
            continue
        # 特別要求：整串『特別要求 Special request：吸菸』顯示在同一個合併格
        if field == "special_request":
            _fill_special_request(ws, lc, v, orig_special_label)
            continue
        cell = _input_cell(ws, lc)
        cell.value = v
        if isinstance(v, (datetime, date)):
            cell.number_format = _CHECK_FMT if field in ("checkin", "checkout") else _DOB_FMT

    # 房型欄不自動勾選（依用戶指示，房型留空由人工選擇）
    # if b.get("room_type"):
    #     _fill_room_type(ws, b["room_type"])


def _safe_sheet_title(base: str, idx: int, surname: str) -> str:
    """產生合法且不重複的 sheet 名（<=31 字、去除非法字元）。"""
    surname = re.sub(r"[\[\]\:\*\?\/\\]", "", surname or "")[:12]
    title = f"{base}-{idx}"
    if surname:
        title = f"{title}-{surname}"
    return title[:31]


# ---------------------------------------------------------------------------
# 清單表填入
# ---------------------------------------------------------------------------
def _col_to_idx(letter: str) -> int:
    return openpyxl.utils.column_index_from_string(letter)


def _clear_list_region(ws, cfg: dict, rows: int = 60):
    """清掉清單表原有的範例資料（這些「空白檔」其實內含範例列）。
    清除範圍涵蓋所有映射欄位並外擴 3 欄（涵蓋吸煙等附加欄），
    自資料起始列往下 rows 列。"""
    cols = cfg["list_cols"]
    idxs = [_col_to_idx(v) for v in cols.values()]
    cmin, cmax = min(idxs), max(idxs) + 3
    start = cfg["list_start_row"]
    for r in range(start, start + rows):
        for c in range(cmin, cmax + 1):
            cell = ws.cell(row=r, column=c)
            mr = _merged_range_of(ws, r, c)
            if mr and (r, c) != (mr.min_row, mr.min_col):
                continue
            cell.value = None


def _nights(ci, co):
    """晚數 = 退房 - 入住（天）。兩者皆為日期才計算。"""
    if isinstance(ci, datetime) and isinstance(co, datetime):
        d = (co - ci).days
        return d if d >= 0 else None
    return None


def _fill_list_sheet(ws, cfg: dict, bookings: list):
    """填工作表1（10 欄訂房摘要表）：每位客人一列。

    自動填：群組 / 代理 / 入住人 / 人數 / 入住日期 / 退房 / 晚數。
    留空白手動補：股東 / 抵澳時間 / 離澳時間。
    所有資料格統一畫上細框線並水平垂直置中。
    """
    _clear_list_region(ws, cfg)
    start = cfg["list_start_row"]
    cols = cfg["list_cols"]

    # 統一樣式：細框線 + 水平垂直置中
    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")

    # 先給整個填表區域（A~J 的資料列）畫上框線、置中
    col_idx = [_col_to_idx(v) for v in cols.values()]
    min_col, max_col = min(col_idx), max(col_idx)
    for i in range(len(bookings)):
        row = start + i
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=c)
            cell.border = border
            cell.alignment = center

    for i, b in enumerate(bookings):
        row = start + i
        # 群組（微信/群組）
        group = b.get("group")
        if group:
            ws.cell(row=row, column=_col_to_idx(cols["group"]), value=group)
        # 代理（訂房人/代理）
        agent = b.get("agent")
        if agent:
            ws.cell(row=row, column=_col_to_idx(cols["agent"]), value=agent)
        # 入住人：中文姓名優先，否則英文
        guest = b.get("zh") or b.get("en") or ""
        if guest:
            ws.cell(row=row, column=_col_to_idx(cols["guest"]), value=guest)
        # 人數（整筆訂房共用）
        pax = b.get("pax")
        if pax not in ("", None):
            ws.cell(row=row, column=_col_to_idx(cols["pax"]), value=pax)
        # 入住日期
        ci = b.get("checkin")
        if ci not in ("", None):
            c = ws.cell(row=row, column=_col_to_idx(cols["checkin"]), value=ci)
            if isinstance(ci, datetime):
                c.number_format = _CHECK_FMT
        # 退房
        co = b.get("checkout")
        if co not in ("", None):
            c = ws.cell(row=row, column=_col_to_idx(cols["checkout"]), value=co)
            if isinstance(co, datetime):
                c.number_format = _CHECK_FMT
        # 晚數
        n = _nights(ci, co)
        if n is not None:
            ws.cell(row=row, column=_col_to_idx(cols["nights"]), value=n)
        # 股東 / 抵澳時間 / 離澳時間 -> 留空白手動補（但框線/置中已套好）


# ---------------------------------------------------------------------------
# 主函式
# ---------------------------------------------------------------------------
def _render(cfg: dict, bookings: list) -> str:
    base_dir = os.path.dirname(os.path.abspath(__file__))
    template_path = os.path.join(base_dir, "templates", cfg["file"])
    wb = openpyxl.load_workbook(template_path)

    # 1) 清單表（若有）填全部客人
    if cfg.get("list_sheet") and cfg["list_sheet"] in wb.sheetnames:
        _fill_list_sheet(wb[cfg["list_sheet"]], cfg, bookings)

    # 2) 每位客人各一張訂房單主表格
    form_name = cfg["form_sheet"]
    src_form = wb[form_name]
    # 從原始（未修改）模板擷取特別要求標籤文字，供填表時做基底（避免複製表累積值）
    _orig_special = None
    slc = _find_label(src_form, HT.FORM_LABELS.get("special_request", []))
    if slc is not None:
        _orig_special = slc.value
    # 同樣記錄原始房型選項文字，複製新表時可正確清空勾選
    _orig_room_cells = {
        (c.row, c.column): c.value
        for c in _find_room_type_cells(src_form)
    }
    if bookings:
        _fill_form_sheet(src_form, bookings[0], _orig_special, _orig_room_cells)
        src_form.title = _safe_sheet_title(form_name, 1, bookings[0].get("surname", ""))
        for idx, b in enumerate(bookings[1:], start=2):
            new_ws = wb.copy_worksheet(src_form)
            _clear_form_values(new_ws, _orig_room_cells)
            _fill_form_sheet(new_ws, b, _orig_special, _orig_room_cells)
            new_ws.title = _safe_sheet_title(form_name, idx, b.get("surname", ""))

    ts = taipei_now().strftime("%Y%m%d_%H%M%S")
    out_name = f"{cfg.get('key', 'booking')}_{ts}.xlsx"
    out_path = os.path.join(tempfile.gettempdir(), out_name)
    _add_agent_order_sheet(wb, bookings)
    wb.save(out_path)
    return out_path


# ---------------------------------------------------------------------------
# 代理訂單分頁
# ---------------------------------------------------------------------------
_AGENT_SHEET_NAME = "代理訂單"
_AGENT_HEADERS = ["代理", "訂單編號", "英文姓名", "中文姓名", "入住日期", "退房日期"]
_AGENT_DISPLAY = {"AT": "信威"}


def _agent_display(code: str) -> str:
    """代理代碼 → 顯示名稱。"""
    if not code:
        return ""
    return _AGENT_DISPLAY.get(code.strip(), code.strip())


def _fmt_date_cell(v):
    """把 checkin/checkout 統一為 date 物件（顯示 2026/07/31）。"""
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        if isinstance(v, datetime):
            return v.date()
        return v
    s = str(v).strip()
    m = re.match(r"(\d{4})[-/](\d{1,2})[-/](\d{1,2})", s)
    if m:
        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
    m = re.search(r"(\d{1,2})[./](\d{1,2})", s)
    if m:
        y = taipei_now().year
        return date(y, int(m.group(1)), int(m.group(2)))
    return None


def _add_agent_order_sheet(wb, bookings):
    """新增『代理訂單』分頁，標題列 + 每位客人一列。

    欄位：代理 | 訂單編號 | 英文姓名 | 中文姓名 | 入住日期 | 退房日期
    """
    # 若已存在同名 sheet，刪除後重建
    if _AGENT_SHEET_NAME in wb.sheetnames:
        del wb[_AGENT_SHEET_NAME]
    ws = wb.create_sheet(_AGENT_SHEET_NAME)
    # 標題列
    for col, h in enumerate(_AGENT_HEADERS, start=1):
        c = ws.cell(row=1, column=col, value=h)
        c.font = c.font.copy(bold=True)
        c.alignment = Alignment(horizontal="center", vertical="center")
    # 資料列
    row = 2
    for b in bookings:
        agent = _agent_display(b.get("agent") or "")
        en_name = b.get("en") or b.get("surname", "")
        if b.get("firstname"):
            en_name = f"{b.get('surname','')},{b.get('firstname','')}".strip(",")
        en_name = (en_name or "").upper()
        zh_name = b.get("zh") or ""
        ci = _fmt_date_cell(b.get("checkin"))
        co = _fmt_date_cell(b.get("checkout"))
        ws.cell(row=row, column=1, value=agent)
        ws.cell(row=row, column=2, value=None)  # 訂單編號留空待人工填/自動產生
        ws.cell(row=row, column=3, value=en_name)
        ws.cell(row=row, column=4, value=zh_name)
        cci = ws.cell(row=row, column=5, value=ci)
        cco = ws.cell(row=row, column=6, value=co)
        if ci:
            cci.number_format = "yyyy/mm/dd"
        if co:
            cco.number_format = "yyyy/mm/dd"
        for col in range(1, 7):
            ws.cell(row=row, column=col).alignment = Alignment(
                horizontal="center", vertical="center"
            )
        row += 1
    # 欄寬
    widths = [10, 14, 22, 14, 14, 14]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w



def fill(hotel_key: str, records: list) -> str:
    """掃描記錄 -> Excel。"""
    if hotel_key not in HT.HOTELS:
        raise ValueError(f"未知酒店代碼：{hotel_key}")
    cfg = dict(HT.HOTELS[hotel_key])
    cfg["key"] = hotel_key
    bookings = [_normalize_record(r) for r in records]
    return _render(cfg, bookings)


def fill_manual(hotel_key: str, booking: dict) -> str:
    """文字訂房 -> Excel。"""
    if hotel_key not in HT.HOTELS:
        raise ValueError(f"未知酒店代碼：{hotel_key}")
    cfg = dict(HT.HOTELS[hotel_key])
    cfg["key"] = hotel_key
    bookings = _bookings_from_manual(booking)
    if not bookings:
        raise ValueError("訂房無客人資料")
    return _render(cfg, bookings)


def _clear_form_values(ws, orig_room_cells=None):
    """清掉複製表中已填的所有主表格欄位值。"""
    for subs in HT.FORM_LABELS.values():
        lc = _find_label(ws, subs)
        if lc is not None:
            _input_cell(ws, lc).value = None
    # 還原房型勾選框為原始未勾選文字
    if orig_room_cells:
        for (row, col), orig_val in orig_room_cells.items():
            ws.cell(row=row, column=col).value = orig_val


if __name__ == "__main__":
    # 簡易自測：掃描假資料填六酒店
    sample_records = [
        {
            "doc_type": "護照",
            "parsed": {
                "last_name": "STEPHENS", "first_name": "JOHN",
                "doc_number": "GBN123456", "date_of_birth": "1985-01-01",
            },
        },
        {
            "doc_type": "護照",
            "parsed": {
                "last_name": "CHAN", "first_name": "TAI MAN",
                "doc_number": "K12345678", "date_of_birth": "1990-06-15",
            },
        },
    ]
    for hk in HT.HOTEL_ORDER:
        p = fill(hk, sample_records)
        print(hk, "->", p, "sheets:", openpyxl.load_workbook(p).sheetnames)

    # 文字訂房自測
    booking = {
        "guests": ["江-泰哥-呂布"],
        "check_in": "2026-07-21", "check_out": "2026-07-23",
        "room_count": 1, "pax": 1, "smoking": True,
    }
    p = fill_manual("londoner", booking)
    wb = openpyxl.load_workbook(p)
    ws = wb["Londoner-1-江"]
    print("文字訂房 倫敦人：", "D16=", ws["D16"].value, "K16=", ws["K16"].value,
          "D20=", ws["D20"].value, "D21=", ws["D21"].value,
          "O20=", ws["O20"].value, "O21=", ws["O21"].value)
